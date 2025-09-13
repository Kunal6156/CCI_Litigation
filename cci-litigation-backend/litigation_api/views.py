from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from django.db.models import Q, Count, Case as DjangoCase, When, IntegerField
from django.contrib.auth import authenticate
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils.html import strip_tags
import datetime
import logging
from datetime import timedelta, date
import json
import csv
import io
import re
from decimal import Decimal
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from datetime import timedelta
import json

from .models import User, Case, Department,NotificationLog
from .serializers import (
    UserSerializer, CaseSerializer, DepartmentSerializer, 
    MyTokenObtainPairSerializer, UserSummarySerializer,
    CaseSummarySerializer
)
from .permissions import IsAdminUser, IsDepartmentalEmployeeOrAdmin

logger = logging.getLogger(__name__)


# views.py - Enhanced Draft Management Views

from .models import Draft
from rest_framework.decorators import action
import json
import uuid

class DraftViewSet(viewsets.ModelViewSet):
    """Enhanced Draft management with persistent storage"""
    
    serializer_class = None  # We'll define this inline
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Get drafts for current user only"""
        return Draft.objects.filter(user=self.request.user)
    
    def get_serializer_class(self):
        """Define serializer inline"""
        from rest_framework import serializers
        
        class DraftSerializer(serializers.ModelSerializer):
            age_in_minutes = serializers.ReadOnlyField()
            is_recent = serializers.ReadOnlyField()
            
            class Meta:
                model = Draft
                fields = [
                    'id', 'draft_key', 'title', 'draft_type', 'form_data',
                    'case_id', 'is_auto_saved', 'created_at', 'updated_at',
                    'age_in_minutes', 'is_recent'
                ]
                read_only_fields = ['id', 'draft_key', 'created_at', 'updated_at']
        
        return DraftSerializer
    
    def perform_create(self, serializer):
        """Set user automatically"""
        serializer.save(user=self.request.user)
    
    @action(detail=False, methods=['post'])
    def auto_save(self, request):
        """Enhanced auto-save with persistent storage"""
        try:
            user = request.user
            case_id = request.data.get('case_id')
            form_data = request.data.get('form_data', {})
            draft_type = request.data.get('draft_type', 'case')
            title = request.data.get('title', '')
            
            # Create unique draft key
            if case_id:
                draft_key = f"{draft_type}_{user.id}_{case_id}_edit"
            else:
                draft_key = f"{draft_type}_{user.id}_new"
            
            # Try to update existing draft or create new one
            draft, created = Draft.objects.update_or_create(
                user=user,
                draft_key=draft_key,
                defaults={
                    'draft_type': draft_type,
                    'title': title or (f"Draft for Case {case_id}" if case_id else "New Case Draft"),
                    'form_data': form_data,
                    'case_id': case_id,
                    'is_auto_saved': True,
                }
            )
            
            # Clean up old auto-saved drafts (keep only latest 10 per user)
            self.cleanup_old_drafts(user)
            
            logger.info(f"Auto-save {'created' if created else 'updated'} for user {user.username}")
            
            return Response({
                'success': True,
                'message': 'Draft saved successfully',
                'draft_key': draft.draft_key,
                'draft_id': draft.id,
                'timestamp': draft.updated_at.isoformat(),
                'created': created
            })
            
        except Exception as e:
            logger.error(f"Auto-save failed for user {request.user.username}: {str(e)}")
            return Response({
                'success': False,
                'error': 'Auto-save failed',
                'details': str(e)
            }, status=500)
    
    @action(detail=False, methods=['get'])
    def get_draft(self, request):
        """Get specific draft by key or case_id"""
        draft_key = request.query_params.get('draft_key')
        case_id = request.query_params.get('case_id')
        draft_type = request.query_params.get('draft_type', 'case')
        
        try:
            if draft_key:
                draft = Draft.objects.get(user=request.user, draft_key=draft_key)
            elif case_id:
                # Try to find draft for editing this case
                draft_key_pattern = f"{draft_type}_{request.user.id}_{case_id}_edit"
                draft = Draft.objects.get(user=request.user, draft_key=draft_key_pattern)
            else:
                # Get the most recent new draft
                draft_key_pattern = f"{draft_type}_{request.user.id}_new"
                draft = Draft.objects.filter(
                    user=request.user, 
                    draft_key__startswith=draft_key_pattern
                ).first()
            
            if draft:
                serializer = self.get_serializer(draft)
                return Response({
                    'success': True,
                    'draft': serializer.data
                })
            else:
                return Response({
                    'success': True,
                    'draft': None
                })
                
        except Draft.DoesNotExist:
            return Response({
                'success': True,
                'draft': None
            })
        except Exception as e:
            logger.error(f"Draft retrieval failed: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to retrieve draft'
            }, status=500)
    
    @action(detail=False, methods=['get'])
    def list_user_drafts(self, request):
        """List all drafts for current user"""
        draft_type = request.query_params.get('draft_type', 'case')
        include_auto_saved = request.query_params.get('include_auto_saved', 'true').lower() == 'true'
        limit = int(request.query_params.get('limit', 20))
        
        queryset = self.get_queryset().filter(draft_type=draft_type)
        
        if not include_auto_saved:
            queryset = queryset.filter(is_auto_saved=False)
        
        queryset = queryset[:limit]
        
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'success': True,
            'drafts': serializer.data,
            'count': queryset.count()
        })
    
    @action(detail=False, methods=['post'])
    def save_manual_draft(self, request):
        """Save a draft manually (user-initiated)"""
        try:
            user = request.user
            title = request.data.get('title', 'Manual Draft')
            form_data = request.data.get('form_data', {})
            draft_type = request.data.get('draft_type', 'case')
            case_id = request.data.get('case_id')
            
            draft = Draft.objects.create(
                user=user,
                title=title,
                draft_type=draft_type,
                form_data=form_data,
                case_id=case_id,
                is_auto_saved=False
            )
            
            serializer = self.get_serializer(draft)
            
            return Response({
                'success': True,
                'message': 'Draft saved successfully',
                'draft': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Manual draft save failed: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to save draft'
            }, status=500)
    
    @action(detail=True, methods=['delete'])
    def delete_draft(self, request, pk=None):
        """Delete a specific draft"""
        try:
            draft = self.get_object()
            draft_title = draft.title
            draft.delete()
            
            return Response({
                'success': True,
                'message': f'Draft "{draft_title}" deleted successfully'
            })
            
        except Exception as e:
            logger.error(f"Draft deletion failed: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to delete draft'
            }, status=500)
    
    @action(detail=False, methods=['delete'], url_path='clear_auto_save')
    def clear_auto_save(self, request):
        """
        Clear a specific auto-saved draft by draft_key
        """
        draft_key = request.query_params.get('draft_key')
        if not draft_key:
            return Response(
                {"success": False, "error": "draft_key query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted, _ = Draft.objects.filter(
            draft_key=draft_key,
            user=request.user
        ).delete()

        if deleted:
            return Response({"success": True, "message": "Draft cleared successfully"})
        return Response(
            {"success": False, "error": "Draft not found"},
            status=status.HTTP_404_NOT_FOUND
        )

    
    @action(detail=False, methods=['delete'])
    def cleanup_drafts(self, request):
        """Clean up old auto-saved drafts"""
        try:
            user = request.user
            days_old = int(request.query_params.get('days_old', 7))
            
            from django.utils import timezone
            from datetime import timedelta
            
            cutoff_date = timezone.now() - timedelta(days=days_old)
            
            # Delete old auto-saved drafts
            deleted_count = Draft.objects.filter(
                user=user,
                is_auto_saved=True,
                updated_at__lt=cutoff_date
            ).delete()[0]
            
            return Response({
                'success': True,
                'message': f'Cleaned up {deleted_count} old drafts'
            })
            
        except Exception as e:
            logger.error(f"Draft cleanup failed: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to clean up drafts'
            }, status=500)
    
    def cleanup_old_drafts(self, user, keep_count=10):
        """Helper method to clean up old auto-saved drafts"""
        try:
            # Get auto-saved drafts older than the keep_count
            old_drafts = Draft.objects.filter(
                user=user,
                is_auto_saved=True
            ).order_by('-updated_at')[keep_count:]
            
            if old_drafts:
                Draft.objects.filter(
                    id__in=[draft.id for draft in old_drafts]
                ).delete()
                
        except Exception as e:
            logger.error(f"Old draft cleanup failed: {str(e)}")


# Default notification templates
DEFAULT_SMS_TEMPLATE = "Dear {advocate_name}, Your case {case_id} has a hearing on {hearing_date} at {court_name}. Please be prepared. - CCI Legal"

DEFAULT_EMAIL_TEMPLATE = """Dear {advocate_name},

This is a reminder for your case {case_id} ({case_type}).

Hearing Details:
- Date: {hearing_date}
- Court/Tribunal: {pending_before_court}
- Petitioner: {party_petitioner}
- Respondent: {party_respondent}
- Financial Implications: {financial_implications}

Please ensure you are well-prepared.

Best regards,
CCI Legal Team
Department: {internal_department}"""


def format_notification_template(template, case):
    """Format notification template with case data"""
    return template.format(
        advocate_name=getattr(case, 'advocate_name', 'Advocate'),
        case_id=case.case_id,
        case_type=getattr(case, 'case_type', 'Legal Case'),
        hearing_date=case.next_hearing_date.strftime('%d-%m-%Y') if getattr(case, 'next_hearing_date', None) else 'TBD',
        party_petitioner=getattr(case, 'parties_involved_complainant', '')[:100] + ('...' if len(getattr(case, 'parties_involved_complainant', '')) > 100 else ''),
        party_respondent=getattr(case, 'parties_involved_opposite', '')[:100] + ('...' if len(getattr(case, 'parties_involved_opposite', '')) > 100 else ''),
        pending_before_court=getattr(case, 'bench', ''),
        financial_implications=getattr(case, 'financial_implications', ''),
        internal_department=getattr(case, 'department_name', ''),
        court_name=getattr(case, 'bench', '')
    )


def send_sms(phone_number, message):
    """Send SMS using your SMS gateway"""
    # For now, just log the SMS - implement actual SMS sending later
    logger.info(f"SMS to {phone_number}: {message}")
    # TODO: Implement SMS sending logic using services like:
    # - Twilio
    # - AWS SNS
    # - MSG91
    # - TextLocal


def send_email(email, subject, content):
    """Send email using Django's email backend"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    try:
        # For development, just log the email - uncomment send_mail for production
        logger.info(f"Email to {email}: {subject}")
        logger.info(f"Content: {content}")
        
        # Uncomment for actual email sending:
        # send_mail(
        #     subject=subject,
        #     message=content,
        #     from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@cci.gov.in'),
        #     recipient_list=[email],
        #     fail_silently=False
        # )
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        raise e


# Notification API Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def upcoming_hearings(request):
    """Get cases with upcoming hearings within specified days"""
    days_ahead = int(request.GET.get('days_ahead', 1))
    department = request.GET.get('department')
    
    today = timezone.now().date()
    target_date = today + timedelta(days=days_ahead)
    
    # Use the correct field name from your Case model
    queryset = Case.objects.filter(
        next_hearing_date__lte=target_date,
        next_hearing_date__gte=today
    )
    
    if not request.user.is_admin and department:
        queryset = queryset.filter(internal_department=department)
    
    serializer = CaseSerializer(queryset, many=True)
    return Response({
        'results': serializer.data,
        'count': queryset.count()
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_hearing_reminders(request):
    """Send SMS/Email reminders for upcoming hearings"""
    data = request.data
    hearing_ids = data.get('hearing_ids')
    days_ahead = data.get('days_ahead', 1)
    departments = data.get('departments', [])
    sms_enabled = data.get('sms_enabled', True)
    email_enabled = data.get('email_enabled', True)
    custom_templates = data.get('custom_templates', {})
    
    if hearing_ids:
        cases = Case.objects.filter(id__in=hearing_ids)
    else:
        today = timezone.now().date()
        target_date = today + timedelta(days=days_ahead)
        cases = Case.objects.filter(
            next_hearing_date__lte=target_date,
            next_hearing_date__gte=today
        )
        if departments:
            cases = cases.filter(internal_department__in=departments)
    
    notifications_sent = 0
    errors = []
    
    for case in cases:
        try:
            # Send SMS
            if sms_enabled and getattr(case, 'advocate_mobile', None):
                sms_content = format_notification_template(
                    custom_templates.get('sms', DEFAULT_SMS_TEMPLATE),
                    case
                )
                send_sms(case.advocate_mobile, sms_content)
                
                # Create actual database record
                NotificationLog.objects.create(
                    case=case,
                    notification_type='sms',
                    recipient=case.advocate_mobile,
                    message_content=sms_content,
                    status='sent',
                    sent_at=timezone.now()
                )
                notifications_sent += 1
            
            # Send Email
            if email_enabled and getattr(case, 'advocate_email', None):
                email_content = format_notification_template(
                    custom_templates.get('email', DEFAULT_EMAIL_TEMPLATE),
                    case
                )
                send_email(case.advocate_email, f"Hearing Reminder - {case.case_id}", email_content)
                
                # Create actual database record
                NotificationLog.objects.create(
                    case=case,
                    notification_type='email',
                    recipient=case.advocate_email,
                    message_content=email_content,
                    status='sent',
                    sent_at=timezone.now()
                )
                notifications_sent += 1
                
        except Exception as e:
            logger.error(f"Failed to send notification for case {case.case_id}: {str(e)}")
            errors.append({
                'case_id': case.case_id,
                'error': str(e)
            })
    
    return Response({
        'notifications_sent': notifications_sent,
        'errors': errors
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_history(request):
    """Get notification history"""
    limit = int(request.GET.get('limit', 50))
    department = request.GET.get('department')
    
    # Query actual NotificationLog records from database
    queryset = NotificationLog.objects.all().order_by('-created_at')
    
    # Filter by department if specified and user is not admin
    if not request.user.is_admin and department:
        queryset = queryset.filter(case__internal_department=department)
    elif not request.user.is_admin:
        # Non-admin users see only their department's notifications
        queryset = queryset.filter(case__internal_department=request.user.department_name)
    
    # Apply limit
    queryset = queryset[:limit]
    
    # Serialize the actual data
    results = []
    for notification in queryset:
        results.append({
            'id': notification.id,
            'case': {
                'case_id': notification.case.case_id if notification.case else 'Unknown',
                'id': notification.case.id if notification.case else None
            },
            'notification_type': notification.notification_type,
            'recipient': notification.recipient,
            'status': notification.status,
            'created_at': notification.created_at,
            'sent_at': notification.sent_at,
            'error_message': notification.error_message
        })
    
    return Response({
        'results': results,
        'count': len(results)
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_manual_notification(request):
    """Send manual notification for a specific case"""
    data = request.data
    case_id = data.get('case_id')
    message_type = data.get('message_type', 'both')
    custom_message = data.get('custom_message', '')
    recipient_phone = data.get('recipient_phone', '')
    recipient_email = data.get('recipient_email', '')
    use_template = data.get('use_template', True)
    
    try:
        case = Case.objects.get(case_id=case_id)
    except Case.DoesNotExist:
        return Response({'error': 'Case not found'}, status=404)
    
    notifications_sent = 0
    errors = []
    
    try:
        if message_type in ['sms', 'both']:
            phone = recipient_phone or getattr(case, 'advocate_mobile', '')
            if phone:
                message = custom_message if not use_template else format_notification_template(DEFAULT_SMS_TEMPLATE, case)
                send_sms(phone, message)
                
                # Create actual database record
                NotificationLog.objects.create(
                    case=case,
                    notification_type='sms',
                    recipient=phone,
                    message_content=message,
                    status='sent',
                    sent_at=timezone.now()
                )
                notifications_sent += 1
        
        if message_type in ['email', 'both']:
            email = recipient_email or getattr(case, 'advocate_email', '')
            if email:
                message = custom_message if not use_template else format_notification_template(DEFAULT_EMAIL_TEMPLATE, case)
                send_email(email, f"Manual Notification - {case.case_id}", message)
                
                # Create actual database record
                NotificationLog.objects.create(
                    case=case,
                    notification_type='email',
                    recipient=email,
                    message_content=message,
                    status='sent',
                    sent_at=timezone.now()
                )
                notifications_sent += 1
        
        return Response({
            'success': True,
            'notifications_sent': notifications_sent,
            'message': 'Manual notification sent successfully'
        })
        
    except Exception as e:
        logger.error(f"Failed to send manual notification: {str(e)}")
        return Response({'error': str(e)}, status=500)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def notification_settings(request):
    """Get or update notification settings"""
    user = request.user
    
    if request.method == 'GET':
        # Return default settings - implement with user preferences model later
        return Response({
            'sms_enabled': True,
            'email_enabled': True,
            'auto_reminders': True,
            'reminder_days_before': 1,
            'notification_time': '09:00',
            'include_departments': [user.department_name] if not user.is_admin else [],
            'sms_template': DEFAULT_SMS_TEMPLATE,
            'email_template': DEFAULT_EMAIL_TEMPLATE
        })
    
    elif request.method == 'POST':
        # Save settings - implement with user preferences model later
        settings = request.data
        logger.info(f"Notification settings updated for user {user.username}: {settings}")
        
        return Response({
            'success': True,
            'message': 'Notification settings saved successfully'
        })

# Add these bulk paste endpoints to your views.py file

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def preview_bulk_paste(request):
    """Preview bulk paste data before actual import"""
    try:
        data = request.data
        paste_data = data.get('paste_data', [])
        
        if not paste_data:
            return Response({
                'error': 'No data provided for preview'
            }, status=400)
        
        # Process and validate the data
        processed_data = []
        errors = []
        valid_rows = 0
        
        for row_index, row in enumerate(paste_data):
            try:
                if len(row) < 3:  # Minimum required columns
                    errors.append(f"Row {row_index + 1}: Insufficient columns (minimum 3 required)")
                    continue
                
                # Map tab-separated data to case fields
                # Adjust this mapping based on your Excel format
                case_data = {
                    'case_type': row[0].strip() if len(row) > 0 else '',
                    'case_number': row[1].strip() if len(row) > 1 else '',
                    'case_year': row[2].strip() if len(row) > 2 else '',
                    'department_name': row[3].strip() if len(row) > 3 else request.user.department_name,
                    'bench': row[4].strip() if len(row) > 4 else '',
                    'parties_involved_complainant': row[5].strip() if len(row) > 5 else '',
                    'parties_involved_opposite': row[6].strip() if len(row) > 6 else '',
                    'financial_implications': row[7].strip() if len(row) > 7 else '0',
                    'status_of_case': row[8].strip() if len(row) > 8 else 'pending',
                }
                
                # Generate case_id
                if case_data['case_type'] and case_data['case_number'] and case_data['case_year']:
                    case_data['case_id'] = f"{case_data['case_type']}/{case_data['case_number']}/{case_data['case_year']}"
                else:
                    errors.append(f"Row {row_index + 1}: Missing case type, number, or year")
                    continue
                
                # Basic validation
                validation_errors = []
                
                if not case_data['case_type']:
                    validation_errors.append("Case type is required")
                
                if not case_data['case_number']:
                    validation_errors.append("Case number is required")
                
                if not case_data['case_year']:
                    validation_errors.append("Case year is required")
                
                # Check for duplicate case_id
                if Case.objects.filter(case_id=case_data['case_id']).exists():
                    validation_errors.append(f"Case ID {case_data['case_id']} already exists")
                
                if validation_errors:
                    errors.append(f"Row {row_index + 1}: {'; '.join(validation_errors)}")
                    case_data['valid'] = False
                else:
                    case_data['valid'] = True
                    valid_rows += 1
                
                processed_data.append(case_data)
                
            except Exception as e:
                errors.append(f"Row {row_index + 1}: Processing error - {str(e)}")
        
        return Response({
            'total_rows': len(paste_data),
            'valid_rows': valid_rows,
            'errors': errors,
            'sample_data': processed_data[:10],  # First 10 rows for preview
            'preview_only': True
        })
        
    except Exception as e:
        logger.error(f"Bulk paste preview failed: {str(e)}")
        return Response({
            'error': 'Preview failed',
            'details': str(e)
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_paste_cases(request):
    """Bulk create cases from paste data"""
    try:
        data = request.data
        paste_data = data.get('paste_data', [])
        options = data.get('options', {})
        skip_duplicates = options.get('skip_duplicates', True)
        validate_only = options.get('validate_only', False)
        
        if not paste_data:
            return Response({
                'error': 'No data provided for import'
            }, status=400)
        
        created_cases = []
        errors = []
        skipped = 0
        
        for row_index, row in enumerate(paste_data):
            try:
                if len(row) < 3:
                    errors.append({
                        'row': row_index + 1,
                        'error': 'Insufficient columns (minimum 3 required)'
                    })
                    continue
                
                # Map data to case fields
                case_data = {
                    'case_type': row[0].strip() if len(row) > 0 else '',
                    'case_number': row[1].strip() if len(row) > 1 else '',
                    'case_year': row[2].strip() if len(row) > 2 else '',
                    'department_name': row[3].strip() if len(row) > 3 else request.user.department_name,
                    'bench': row[4].strip() if len(row) > 4 else '',
                    'parties_involved_complainant': row[5].strip() if len(row) > 5 else '',
                    'parties_involved_opposite': row[6].strip() if len(row) > 6 else '',
                    'financial_implications': row[7].strip() if len(row) > 7 else '0',
                    'status_of_case': row[8].strip() if len(row) > 8 else 'pending',
                    'date_of_filing': timezone.now().date(),  # Default to today
                    'created_by': request.user,
                    'last_updated_by': request.user,
                }
                
                # Generate case_id
                if case_data['case_type'] and case_data['case_number'] and case_data['case_year']:
                    case_data['case_id'] = f"{case_data['case_type']}/{case_data['case_number']}/{case_data['case_year']}"
                else:
                    errors.append({
                        'row': row_index + 1,
                        'error': 'Missing case type, number, or year'
                    })
                    continue
                
                # Check for duplicates
                if Case.objects.filter(case_id=case_data['case_id']).exists():
                    if skip_duplicates:
                        skipped += 1
                        continue
                    else:
                        errors.append({
                            'row': row_index + 1,
                            'error': f'Case ID {case_data["case_id"]} already exists'
                        })
                        continue
                
                # Validate only mode
                if validate_only:
                    created_cases.append(case_data)
                    continue
                
                # Create the case
                case_serializer = CaseSerializer(data=case_data)
                if case_serializer.is_valid():
                    case = case_serializer.save()
                    created_cases.append({
                        'case_id': case.case_id,
                        'id': case.id,
                        'row': row_index + 1
                    })
                    logger.info(f"Bulk paste: Created case {case.case_id}")
                else:
                    errors.append({
                        'row': row_index + 1,
                        'error': f'Validation failed: {case_serializer.errors}'
                    })
                
            except Exception as e:
                logger.error(f"Bulk paste error on row {row_index + 1}: {str(e)}")
                errors.append({
                    'row': row_index + 1,
                    'error': str(e)
                })
        
        return Response({
            'created': len(created_cases),
            'errors': errors,
            'skipped': skipped,
            'validate_only': validate_only,
            'created_cases': created_cases if validate_only else []
        })
        
    except Exception as e:
        logger.error(f"Bulk paste failed: {str(e)}")
        return Response({
            'error': 'Bulk paste failed',
            'details': str(e)
        }, status=500)


# Add these URL patterns to your urls.py file
from django.urls import path, include
from rest_framework.routers import DefaultRouter
from . import views

# Add these to your existing urlpatterns:
urlpatterns = [
    # ... your existing patterns ...
    
    # Bulk paste endpoints
    path('cases/preview-bulk-paste/', views.preview_bulk_paste, name='preview-bulk-paste'),
    path('cases/bulk-paste/', views.bulk_paste_cases, name='bulk-paste-cases'),
]

class MyTokenObtainPairView(TokenObtainPairView):
    """
    Enhanced JWT token view with login tracking and department validation.
    """
    serializer_class = MyTokenObtainPairSerializer
    
    def post(self, request, *args, **kwargs):
        """Override to add login tracking and enhanced validation"""
        response = super().post(request, *args, **kwargs)
        
        if response.status_code == 200:
            # Get user info from serializer
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                user = serializer.user
                if user:
                    # Update last login
                    user.last_login = timezone.now()
                    user.save(update_fields=['last_login'])
                    
                    # Log successful login
                    logger.info(f"User {user.username} ({user.department_name}) logged in successfully")
                    
                    # Add user info to response
                    response.data['user_info'] = {
                        'id': user.id,
                        'username': user.username,
                        'email': user.email,
                        'full_name': user.get_full_name(),
                        'department_name': user.department_name,
                        'user_type': user.user_type,
                        'is_admin': user.is_admin,
                        'permissions': {
                            'can_manage_users': user.is_admin,
                            'can_export_data': True,
                            'can_view_all_cases': True,
                            'can_edit_all_cases': user.is_admin,
                            'accessible_departments': [user.department_name] if not user.is_admin else 'all'
                        }
                    }
        else:
            # Log failed login attempt
            username = request.data.get('username', 'Unknown')
            logger.warning(f"Failed login attempt for username: {username}")
        
        return response


class UserViewSet(viewsets.ModelViewSet):
    """
    Enhanced User management viewset with comprehensive role-based access.
    Only accessible by admin users.
    """
    queryset = User.objects.all().order_by('department_name', 'username')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get_queryset(self):
        """Filter users based on department and search"""
        queryset = self.queryset
        
        # Search functionality
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(department_name__icontains=search) |
                Q(employee_id__icontains=search)
            )
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(internal_department=department)
        
        # Filter by user type
        user_type = self.request.query_params.get('user_type', None)
        if user_type and user_type in ['admin', 'user']:
            queryset = queryset.filter(user_type=user_type)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action in ['list', 'summary']:
            return UserSummarySerializer
        return UserSerializer
    
    def create(self, request, *args, **kwargs):
        """Enhanced user creation with validation"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Additional validation for admin creation
        department_name = serializer.validated_data.get('department_name')
        user_type = serializer.validated_data.get('user_type', 'user')
        password = serializer.validated_data.get('password')
        
        if password:
            from django.contrib.auth.password_validation import validate_password
            from django.core.exceptions import ValidationError
            try:
                validate_password(password)
            except ValidationError as e:
                return Response(
                {'password': list(e.messages)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check admin limits per department (2 admins max in Corporate Office)
        if user_type == 'admin' and department_name == 'Corporate Office':
            existing_admins = User.objects.filter(
                department_name='Corporate Office', 
                user_type='admin', 
                is_active=True
            ).count()
            if existing_admins >= 2:
                return Response(
                    {'error': 'Maximum 2 administrators allowed in Corporate Office.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Create user
        user = serializer.save()
        
        # Log user creation
        logger.info(f"User {user.username} created by admin {request.user.username}")
        
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                'message': 'User created successfully',
                'user': UserSummarySerializer(user).data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    def update(self, request, *args, **kwargs):
        """Enhanced user update with validation"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Prevent self-deactivation
        if instance == request.user and 'is_active' in request.data:
            if not request.data['is_active']:
                return Response(
                    {'error': 'You cannot deactivate your own account.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        # Log user update
        logger.info(f"User {instance.username} updated by admin {request.user.username}")
        
        self.perform_update(serializer)
        
        return Response({
            'message': 'User updated successfully',
            'user': UserSummarySerializer(instance).data
        })
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete user (deactivate instead of delete)"""
        instance = self.get_object()
        
        # Prevent self-deletion
        if instance == request.user:
            return Response(
                {'error': 'You cannot delete your own account.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Deactivate instead of delete
        instance.is_active = False
        instance.save()
        
        # Log user deactivation
        logger.info(f"User {instance.username} deactivated by admin {request.user.username}")
        
        return Response(
            {'message': 'User deactivated successfully'},
            status=status.HTTP_204_NO_CONTENT
        )
    
    @action(detail=False, methods=['get', 'put'], permission_classes=[IsAuthenticated])
    def profile(self, request):
        """Get or update user profile"""
        if request.method == 'GET':
            serializer = UserSerializer(request.user)
            return Response(serializer.data)
        elif request.method == 'PUT':
            serializer = UserSerializer(request.user, data=request.data, partial=True)
            if serializer.is_valid():
                protected_fields = ['username', 'is_admin', 'user_type']
                for field in protected_fields:
                    if field in serializer.validated_data:
                        if not request.user.is_admin:
                            return Response(
                            {'error': f'You cannot modify {field}.'},
                            status=status.HTTP_403_FORBIDDEN
                        )
                serializer.save()
                return Response({
                'message': 'Profile updated successfully',
                'user': serializer.data
            })
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['put'], permission_classes=[IsAuthenticated])
    def update_profile(self, request):
        """Update user profile (alias for profile PUT)"""
        return self.profile(request)

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def change_password(self, request):
        """Change user password"""
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        if not old_password or not new_password:
            return Response({'error': 'Both old and new passwords are required.'},status=status.HTTP_400_BAD_REQUEST)
        if not user.check_password(old_password):
            return Response({'error': 'Current password is incorrect.'},status=status.HTTP_400_BAD_REQUEST)
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response({'error': list(e.messages)}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(new_password)
        user.save()
        logger.info(f"Password changed for user {user.username}")
        return Response({'message': 'Password changed successfully'})
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def logout(self, request):
        """Logout user (optional server-side cleanup)"""
        # You can add server-side logout logic here if needed
        # For JWT, the token management is typically handled client-side
        logger.info(f"User {request.user.username} logged out")
        
        return Response({'message': 'Logged out successfully'})
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get user summary statistics"""
        queryset = self.get_queryset()
        
        # Department-wise stats
        dept_stats = {}
        for dept_choice in User.DEPARTMENT_CHOICES:
            dept_name = dept_choice[0]
            dept_users = queryset.filter(department_name=dept_name)
            dept_stats[dept_name] = {
                'total': dept_users.count(),
                'active': dept_users.filter(is_active=True).count(),
                'admins': dept_users.filter(user_type='admin', is_active=True).count(),
                'users': dept_users.filter(user_type='user', is_active=True).count(),
            }
        
        return Response({
            'total_users': queryset.count(),
            'active_users': queryset.filter(is_active=True).count(),
            'total_admins': queryset.filter(user_type='admin', is_active=True).count(),
            'total_regular_users': queryset.filter(user_type='user', is_active=True).count(),
            'department_stats': dept_stats
        })
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Reset user password (admin only)"""
        user = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password:
            return Response(
                {'error': 'New password is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate password
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            return Response(
                {'error': list(e.messages)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Set new password
        user.set_password(new_password)
        user.save()
        
        # Log password reset
        logger.info(f"Password reset for user {user.username} by admin {request.user.username}")
        
        return Response({'message': 'Password reset successfully'})


class DepartmentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Enhanced Department viewset with statistics.
    Accessible by any authenticated user.
    """
    queryset = Department.objects.all().order_by('name')
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def choices(self, request):
        """Get department choices for dropdowns"""
        choices = [{'value': choice[0], 'label': choice[1]} for choice in User.DEPARTMENT_CHOICES]
        return Response(choices)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get department-wise statistics"""
        stats = {}
        
        for dept_choice in User.DEPARTMENT_CHOICES:
            dept_name = dept_choice[0]
            
            # User stats
            users = User.objects.filter(department_name=dept_name, is_active=True)
            
            # Case stats
            cases = Case.objects.filter(department_name=dept_name)
            
            stats[dept_name] = {
                'users': {
                    'total': users.count(),
                    'admins': users.filter(user_type='admin').count(),
                    'regular': users.filter(user_type='user').count(),
                },
                'cases': {
                    'total': cases.count(),
                    'pending': cases.filter(status_of_case__in=['pending', 'under_hearing']).count(),
                    'disposed': cases.filter(status_of_case__in=['disposed', 'closed']).count(),
                }
            }
        
        return Response(stats)


class CaseViewSet(viewsets.ModelViewSet):
    """
    Enhanced Case management viewset with Excel format compatibility, auto-save, and search.
    """
    queryset = Case.objects.all().select_related('created_by', 'last_updated_by')
    serializer_class = CaseSerializer
    permission_classes = [IsAuthenticated, IsDepartmentalEmployeeOrAdmin]
    
    def get_queryset(self):
        """Enhanced filtering based on user role and permissions"""
        user = self.request.user
        queryset = self.queryset
        
        # Role-based filtering
        if not user.is_admin:
            # Regular users can view all cases but edit only their department's cases
            # For viewing, we allow all cases as per requirement
            pass  # Allow viewing all cases
        
        # Enhanced search functionality - Excel format compatible
        search_query = self.request.query_params.get('search', None)
        if search_query:
            # Split search query for better matching
            search_terms = search_query.split()
            search_q = Q()
            
            for term in search_terms:
                search_q |= (
                    Q(case_id__icontains=term) |
                    Q(case_name__icontains=term) |
                    Q(parties_involved_complainant__icontains=term) |
                    Q(parties_involved_opposite__icontains=term) |
                    Q(bench__icontains=term) |
                    Q(sections_involved__icontains=term) |
                    Q(status_of_case__icontains=term) |
                    Q(type_of_litigation__icontains=term) |
                    Q(relief_orders_prayed__icontains=term) |
                    Q(important_directions_orders__icontains=term) |
                    Q(outcome__icontains=term) |
                    Q(department_name__icontains=term)
                )
            
            queryset = queryset.filter(search_q)
        
        # Department filter
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(internal_department=department)
        
        # Status filter
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status_of_case=status_filter)
        
        # Date range filters
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        
        if date_from:
            try:
                from datetime import datetime
                date_from = datetime.strptime(date_from, '%d-%m-%Y').date()
                queryset = queryset.filter(date_of_institution__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to = datetime.strptime(date_to, '%d-%m-%Y').date()
                queryset = queryset.filter(date_of_institution__lte=date_to)
            except ValueError:
                pass
        
        # Sorting
        sort_by = self.request.query_params.get('sort_by', 'date_of_filing')
        sort_order = self.request.query_params.get('sort_order', 'desc')

        # Map frontend field names to actual model fields
        SORT_FIELD_MAP = {
    "date_of_institution": "date_of_filing",  # frontend → backend mapping
        }

        # Apply mapping if necessary
        sort_by = SORT_FIELD_MAP.get(sort_by, sort_by)

        # Validate sort field
        allowed_sort_fields = [f.name for f in Case._meta.get_fields()]
        if sort_by not in allowed_sort_fields:
             sort_by = 'date_of_filing'

        # Apply sort order
        if sort_order == 'desc':
             sort_by = f'-{sort_by}'

        try:
            queryset = queryset.order_by(sort_by)
        except Exception as e:
            logger.warning(f"Invalid sort_by '{sort_by}', defaulting to date_of_filing. Error: {e}")
            queryset = queryset.order_by('-date_of_filing')
        return queryset

    
    def get_serializer_class(self):
        if self.action in ['summary', 'dashboard_stats']:
            return CaseSummarySerializer
        return CaseSerializer
    
    def perform_create(self, serializer):
        """Set creator and department automatically"""
        user = self.request.user
        
        # Non-admin users can only create cases for their department
        if not user.is_admin:
            serializer.save(
                created_by=user,
                last_updated_by=user,
                department_name=user.department_name
            )
        else:
            # Admins can create for any department
            serializer.save(
                created_by=user,
                last_updated_by=user
            )
        
        # Log case creation
        logger.info(f"Case {serializer.instance.case_id} created by {user.username}")
    
    def perform_update(self, serializer):
        """Enhanced update with role-based restrictions"""
        user = self.request.user
        instance = serializer.instance
        
        # Non-admin users can only edit their department's cases
        if not user.is_admin and instance.department_name != user.department_name:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only edit cases from your department.")
        
        # Prevent department change by non-admins
        if not user.is_admin and 'department_name' in serializer.validated_data:
            if serializer.validated_data['department_name'] != instance.department_name:
                from rest_framework import serializers as rf_serializers
                raise rf_serializers.ValidationError({
                    "department_name": "You cannot change the department of a case."
                })
        
        serializer.save(last_updated_by=user)
        
        # Log case update
        logger.info(f"Case {instance.case_id} updated by {user.username}")
    
    def destroy(self, request, *args, **kwargs):
        """Enhanced delete with role-based restrictions"""
        instance = self.get_object()
        user = request.user
        
        # Non-admin users can only delete their department's cases
        if not user.is_admin and instance.department_name != user.department_name:
            return Response(
                {'error': 'You can only delete cases from your department.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Log case deletion
        logger.info(f"Case {instance.case_id} deleted by {user.username}")
        
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def auto_save(self, request):
        """
        Auto-save functionality for accidental shutdown protection.
        Saves draft data without full validation.
        """
        try:
            case_id = request.data.get('case_id')
            user = request.user
            form_data = request.data
            
            if case_id:
              draft_key = f"case_{user.id}_{case_id}_edit"
              title = f"Draft for Case {case_id}"
              
            else:
              draft_key = f"case_{user.id}_new"
              title = "New Case Draft"
              
            draft, created = Draft.objects.update_or_create(
            user=user,
            draft_key=draft_key,
            defaults={
                'draft_type': 'case',
                'title': title,
                'form_data': form_data,
                'case_id': case_id,
                'is_auto_saved': True,})
            
            logger.info(f"Auto-save {'created' if created else 'updated'} for user {user.username}")
            return Response({
            'success': True,
            'message': 'Draft saved successfully',
            'draft_key': draft.draft_key,
            'draft_id': draft.id,
            'timestamp': draft.updated_at.isoformat()})
            
        except Exception as e:
            logger.error(f"Auto-save failed for user {request.user.username}: {str(e)}")
            return Response({
            'success': False,
            'error': 'Auto-save failed'
        }, status=500)
                
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def get_auto_save(self, request):
        """
        Retrieve auto-saved draft data.
        """
        try:
            draft_key = request.query_params.get('draft_key')
            case_id = request.query_params.get('case_id')
            user = request.user
            
            draft=None
            
            if draft_key:
                try:
                    draft = Draft.objects.get(user=user, draft_key=draft_key)
                except Draft.DoesNotExist:
                   pass
               
            elif case_id:
                try:
                    draft_key_pattern = f"case_{user.id}_{case_id}_edit"
                    draft = Draft.objects.get(user=user, draft_key=draft_key_pattern)
                except Draft.DoesNotExist:
                    pass
                
            else:
                draft = Draft.objects.filter(
                user=user,
                draft_type='case',
                case_id__isnull=True
            ).order_by('-updated_at').first()
                
            if draft:
                return Response({
                'success': True,
                'data': draft.form_data,
                'timestamp': draft.updated_at.isoformat(),
                'draft_key': draft.draft_key,
                'title': draft.title
            })
            else:
                return Response({
                'success': True,
                'data': None
            })
                
        except Exception as e:
            logger.error(f"Auto-save retrieval failed: {str(e)}")
            return Response({
            'success': True,
            'data': None
        })
            
    
    @action(detail=False, methods=['delete'], permission_classes=[IsAuthenticated])
    def clear_auto_save(self, request):
        """
        Clear auto-saved draft data.
        """
        try:
            draft_key = request.query_params.get('draft_key')
            case_id = request.query_params.get('case_id')
            user = request.user
            
            if draft_key:
                Draft.objects.filter(user=user, draft_key=draft_key).delete()
            elif case_id:
                draft_key_pattern = f"case_{user.id}_{case_id}_edit"
                Draft.objects.filter(user=user, draft_key=draft_key_pattern).delete()
            else:
                Draft.objects.filter(
                user=user,
                draft_type='case',
                case_id__isnull=True,
                is_auto_saved=True
            ).delete()
            return Response({'success': True})
        
        except Exception as e:
            logger.error(f"Auto-save clear failed: {str(e)}")
            return Response({'success': False})
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get dashboard statistics based on user role"""
        user = request.user
        
        # Base queryset based on user role
        if user.is_admin:
            cases_queryset = Case.objects.all()
        else:
            # Regular users see all cases for viewing but stats can be customized
            cases_queryset = Case.objects.all()
        
        # Calculate statistics
        total_cases = cases_queryset.count()
        pending_cases = cases_queryset.filter(
            status_of_case__in=['pending', 'under_hearing', 'admitted']
        ).count()
        disposed_cases = cases_queryset.filter(
            status_of_case__in=['disposed', 'closed', 'dismissed']
        ).count()
        
        # Overdue cases (where next hearing date has passed)
        overdue_cases = cases_queryset.filter(
            date_of_next_hearing_order__lt=timezone.now().date(),
            status_of_case__in=['pending', 'under_hearing', 'admitted']
        ).count()
        
        # Recent cases (last 30 days)
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        recent_cases = cases_queryset.filter(
            created_at__gte=thirty_days_ago
        ).order_by('-created_at')[:10]
        
        # Department-wise statistics
        dept_stats = {}
        for dept_choice in User.DEPARTMENT_CHOICES:
            dept_name = dept_choice[0]
            dept_cases = cases_queryset.filter(department_name=dept_name)
            dept_stats[dept_name] = {
                'total': dept_cases.count(),
                'pending': dept_cases.filter(present_status__in=['pending', 'under_hearing']).count(),
                'disposed': dept_cases.filter(present_status__in=['disposed', 'closed']).count(),
                'overdue': dept_cases.filter(
                    date_of_next_hearing_order__lt=timezone.now().date(),
                    present_status__in=['pending', 'under_hearing']
                ).count()
            }
        
        # Monthly statistics (last 12 months)
        monthly_stats = {}
        for i in range(12):
            month_date = timezone.now().date().replace(day=1) - timedelta(days=30*i)
            month_key = month_date.strftime('%Y-%m')
            monthly_cases = cases_queryset.filter(
                date_of_institution__year=month_date.year,
                date_of_institution__month=month_date.month
            ).count()
            monthly_stats[month_key] = monthly_cases
        
        return Response({
            'total_cases': total_cases,
            'pending_cases': pending_cases,
            'disposed_cases': disposed_cases,
            'overdue_cases': overdue_cases,
            'recent_cases': CaseSummarySerializer(recent_cases, many=True).data,
            'department_wise_stats': dept_stats,
            'monthly_stats': monthly_stats,
            'user_department': user.department_name,
            'user_permissions': {
                'can_create': True,
                'can_edit_all': user.is_admin,
                'can_delete': user.is_admin or user.user_type == 'admin',
                'can_export': True,
            }
        })
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request, *args, **kwargs):
        """
        Enhanced Excel export matching exact format requirements with Indian formatting.
        """
        # Use the same filtering as get_queryset but ensure role-based access
        queryset = self.get_queryset()
        
        # Additional filtering for export (if needed)
        user = request.user
        if not user.is_admin:
            # For export, regular users get all data as per requirement
            # but this can be restricted if needed
            pass
        
        # Prepare HTTP response for Excel file
        current_time = timezone.now()
        file_name = f"CCI_Litigation_Cases_{current_time.strftime('%d-%m-%Y_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        
        

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "CCI Litigation Cases"
        
        # Define headers matching Excel format exactly
        # First header row (merged titles)
        header_row_1 = [
        "Case Number",           # Col 1 (will merge A1:C1)
        "",                      # Col 2 (part of Case Number merge)
        "",                      # Col 3 (part of Case Number merge)
        "Date of Filing/Intimation",  # Col 4
        "Pending Before Court",  # Col 5
        "Party Details",         # Col 6 (will merge F1:G1)
        "",                      # Col 7 (part of Party Details merge)
        "Nature of Claim",       # Col 8
        "Advocate",              # Col 9 (will merge I1:K1)
        "",                      # Col 10 (part of Advocate merge)
        "",                      # Col 11 (part of Advocate merge)
        "Financial Implications", # Col 12
        "Internal Department of CCI", # Col 13
        "Last Date of Hearing",  # Col 14
        "Next Date of Hearing",  # Col 15
        "Brief Description of Matter", # Col 16
        "Relief Claimed by Party", # Col 17
        "Present Status",        # Col 18
        "Remarks"               # Col 19
    ]
        
        header_row_2 = [
    "Type",        # A2
    "Number",      # B2
    "Year",        # C2
    None,          # D2 - merged from row 1
    None,          # E2 - merged from row 1
    "Petitioner",  # F2
    "Respondent",  # G2
    None,          # H2 - merged from row 1
    "Name",        # I2
    "Email",       # J2
    "Mobile",      # K2
    None,          # L2 - merged from row 1
    None,          # M2 - merged from row 1
    None,          # N2 - merged from row 1
    None,          # O2 - merged from row 1
    None,          # P2 - merged from row 1
    None,          # Q2 - merged from row 1
    None,          # R2 - merged from row 1
    None           # S2 - merged from row 1
]

        ws.append(header_row_1)
        ws.append(header_row_2)
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        # Apply styling ONLY to row 1 (merged header row)
        for col_num in range(1, 20):  # 19 columns total
           cell = ws.cell(row=1, column=col_num)
           cell.font = header_font
           cell.fill = header_fill
           cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
           cell.border = header_border

# Apply minimal styling to row 2 (second header row)
        for col_num in range(1, 20):
            cell = ws.cell(row=2, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = header_border  # optional: keep borders consistent
        
        ws.row_dimensions[2].height = 25


        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)   # Case Number (A1:C1)
       # ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)   # Date of Filing (D1:D2)
        #ws.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)   # Pending Before Court (E1:E2)
        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)   # Party Details (F1:G1)
        #ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=8)   # Nature of Claim (H1:H2)
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)  # Advocate (I1:K1)
        # ws.merge_cells(start_row=1, start_column=12, end_row=2, end_column=12) # Financial (L1:L2)
        # ws.merge_cells(start_row=1, start_column=13, end_row=2, end_column=13) # Internal Dept (M1:M2)
        # ws.merge_cells(start_row=1, start_column=14, end_row=2, end_column=14) # Last Hearing (N1:N2)
        # ws.merge_cells(start_row=1, start_column=15, end_row=2, end_column=15) # Next Hearing (O1:O2)
        # ws.merge_cells(start_row=1, start_column=16, end_row=2, end_column=16) # Brief Desc (P1:P2)
        # ws.merge_cells(start_row=1, start_column=17, end_row=2, end_column=17) # Relief Claimed (Q1:Q2)
        # ws.merge_cells(start_row=1, start_column=18, end_row=2, end_column=18) # Present Status (R1:R2)
        # ws.merge_cells(start_row=1, start_column=19, end_row=2, end_column=19) # Remarks (S1:S2)

# Merge other headers if needed in future

        # Enhanced header styling
        
        merged_ranges = [
        (1, 1, 1, 3),   # Case Number
        (1, 4, 2, 4),   # Date of Filing
        (1, 5, 2, 5),   # Pending Before Court
        (1, 6, 1, 7),   # Party Details
        (1, 8, 2, 8),   # Nature of Claim
        (1, 9, 1, 11),  # Advocate
        (1, 12, 2, 12), # Financial
        (1, 13, 2, 13), # Internal Dept
        (1, 14, 2, 14), # Last Hearing
        (1, 15, 2, 15), # Next Hearing
        (1, 16, 2, 16), # Brief Desc
        (1, 17, 2, 17), # Relief Claimed
        (1, 18, 2, 18), # Present Status
        (1, 19, 2, 19)  # Remarks
    ]
        
        
        for start_row, start_col, end_row, end_col in merged_ranges:
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                    else:
                        cell.font = Font(bold=False)
                        cell.fill = PatternFill(fill_type=None)

                    cell.alignment = header_alignment
                    cell.border = header_border
                
        ws.freeze_panes = 'A3'
            
        

        
        # Add data rows with proper formatting
        for case in queryset:
            # Format case data according to Excel requirements
            case_number = self._format_case_number(case)
            filing_date = self._format_indian_date(case.date_of_filing)
            party_details = self._combine_party_details(case)
            nature_of_claim = getattr(case, 'nature_of_claim', 'Others')  # New field
            advocate_details = self._format_advocate_details(case)  # New field
            financial = self._format_indian_currency(case.financial_implications)
            last_hearing = self._format_indian_date(getattr(case, 'last_hearing_date', None))  # New field
            next_hearing = self._format_indian_date(case.next_hearing_date)
            brief_description = strip_tags(getattr(case, 'brief_description', '') or '')
            relief_claimed = strip_tags(getattr(case, 'relief_claimed', '') or '')  # New field
            present_status = strip_tags(getattr(case, 'present_status', '') or '')# New field
            remarks = strip_tags(getattr(case, 'case_remarks', '') or '') # New field
            
            row = [
    case.case_type or '',
    case.case_number or '',
    case.case_year or '',
    filing_date,
    getattr(case, 'pending_before_court', '') or '',
    self._format_party_list(case.party_petitioner),
    self._format_party_list(case.party_respondent),
    nature_of_claim,
    case.advocate_name or '',
    case.advocate_email or '',
    case.advocate_mobile or '',
    financial,
    case.internal_department or '',
    last_hearing,
    next_hearing,
    brief_description,
    relief_claimed,
    present_status,
    remarks
]

            ws.append(row)
            last_row_idx = ws.max_row
            for col_idx in range(1, 20):  # Adjust the range if more columns are added
                cell = ws.cell(row=last_row_idx, column=col_idx)
                if col_idx in [1, 2, 3]:
                    try:
                        cell.alignment = Alignment(vertical='bottom', horizontal='left', wrap_text=True)
                    except ValueError:
                # Fallback for older openpyxl versions
                        cell.alignment = Alignment(vertical='justify', horizontal='left', wrap_text=True)

                    #cell.alignment = Alignment(vertical='bottom', wrap_text=True)

        
                # Wrap text in Petitioner and Respondent cells
            
            #for cell in ws.cell(row=last_row_idx, column=4), ws.cell(row=last_row_idx, column=5):
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            for col_idx in [6, 7]: 
                cell = ws.cell(row=last_row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        # Auto-adjust column widths with better sizing
        column_widths = [12, 10, 10, 18, 25, 30, 30, 20, 25, 30, 15, 20, 25, 18, 18, 50, 40, 30, 40]

        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = width
        
        # Add data validation and formatting
        self._apply_excel_formatting(ws, queryset.count())
        
        # Add summary sheet for admins
        if user.is_admin:
            summary_ws = wb.create_sheet("Summary Statistics")
            self._create_summary_sheet(summary_ws, queryset, user)
        
        # Add export metadata
        metadata_ws = wb.create_sheet("Export Information")
        self._create_metadata_sheet(metadata_ws, user, queryset.count())
        
        # Save and return
        wb.save(response)
        
        # Log export
        logger.info(f"Excel export generated by {user.username} with {queryset.count()} records")
        
        return response
    
    def _format_case_number(self, case):
        """Format case number as Type/Number/Year"""
        case_type = getattr(case, 'case_type', '')
        case_number = getattr(case, 'case_number', '')
        case_year = getattr(case, 'case_year', '')
        
        if case_type and case_number and case_year:
            return f"{case_type}/{case_number}/{case_year}"
        else:
            return case.case_id or ''
    def _format_party_list(self, party_str):
            if not party_str:
               return ''
            lines = [f"{i+1}. {line.strip()}" for i, line in enumerate(party_str.split(','))]
            return "\n".join(lines)
    
    def _format_indian_date(self, date_obj):
        """Format date as DD-MM-YYYY (Indian format)"""
        if date_obj:
            return date_obj.strftime('%d-%m-%Y')
        return ''
    
    def _combine_party_details(self, case):
        """Combine party details in proper format"""
        complainant = getattr(case, 'parties_involved_complainant', '')
        respondent = getattr(case, 'parties_involved_opposite', '')
        
        if complainant and respondent:
            return f"Petitioner: {complainant}\nRespondent: {respondent}"
        elif complainant:
            return f"Petitioner: {complainant}"
        elif respondent:
            return f"Respondent: {respondent}"
        return ''
    
    def _format_advocate_details(self, case):
        """Format advocate details as Name/Email/Mobile"""
        advocate_name = getattr(case, 'advocate_name', '')
        advocate_email = getattr(case, 'advocate_email', '')
        advocate_mobile = getattr(case, 'advocate_mobile', '')
        
        details = []
        if advocate_name:
            details.append(f"Name: {advocate_name}")
        if advocate_email:
            details.append(f"Email: {advocate_email}")
        if advocate_mobile:
            details.append(f"Mobile: {advocate_mobile}")
        
        return '\n'.join(details)
    
    def _format_indian_currency(self, amount):
        """Format currency in Indian format Rs. 1,00,00,00,00,00,000.00"""
        if not amount:
            return ''
        
        try:
            # Convert to Decimal for precision
            if isinstance(amount, str):
                amount = Decimal(amount)
            elif not isinstance(amount, Decimal):
                amount = Decimal(str(amount))
            
            # Format with Indian numbering system
            amount_str = f"{amount:,.2f}"
            
            # Convert to Indian numbering (add commas at appropriate places)
            parts = amount_str.split('.')
            integer_part = parts[0].replace(',', '')
            decimal_part = parts[1] if len(parts) > 1 else '00'
            
            # Indian numbering system
            if len(integer_part) > 3:
                # First group of 3 from right
                result = integer_part[-3:]
                integer_part = integer_part[:-3]
                
                # Then groups of 2
                while len(integer_part) > 2:
                    result = integer_part[-2:] + ',' + result
                    integer_part = integer_part[:-2]
                
                if integer_part:
                    result = integer_part + ',' + result
            else:
                result = integer_part
            
            return f"Rs. {result}.{decimal_part}"
            
        except (ValueError, TypeError):
            return str(amount) if amount else ''
    
    def _apply_excel_formatting(self, ws, row_count):
        """Apply formatting to Excel sheet"""
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=row_count + 1):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, ws, queryset, user):
        """Create summary statistics sheet"""
        ws.append(["CCI Litigation System - Summary Statistics"])
        ws.append([])
        ws.append(["Department", "Total Cases", "Pending Cases", "Disposed Cases", "Overdue Cases"])
        
        # Department-wise summary
        for dept_choice in User.DEPARTMENT_CHOICES:
            dept_name = dept_choice[0]
            dept_cases = queryset.filter(internal_department=dept_name)
            pending = dept_cases.filter(present_status__in=['pending', 'under_hearing']).count()
            disposed = dept_cases.filter(present_status__in=['disposed', 'closed']).count()

            overdue = dept_cases.filter(
                next_hearing_date__lt=timezone.now().date(),
                present_status__in=['pending', 'under_hearing']
            ).count()
            
            ws.append([dept_name, dept_cases.count(), pending, disposed, overdue])
        
        # Style summary headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _create_metadata_sheet(self, ws, user, record_count):
        """Create export metadata sheet"""
        current_time = timezone.now()
        
        ws.append(["CCI Litigation System - Export Information"])
        ws.append([])
        ws.append(["Export Details", ""])
        ws.append(["Exported By", user.get_full_name()])
        ws.append(["Username", user.username])
        ws.append(["Department", user.department_name])
        ws.append(["User Role", "Administrator" if user.is_admin else "Regular User"])
        ws.append(["Export Date", current_time.strftime('%d-%m-%Y')])
        ws.append(["Export Time", current_time.strftime('%H:%M:%S')])
        ws.append(["Total Records", record_count])
        ws.append(["File Format", "Excel (.xlsx)"])
        ws.append(["System Version", "CCI Litigation v2.0"])
        
        # Style metadata
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=12), 3):
            row[0].font = Font(bold=True)


# Additional utility views for Excel format compatibility
@method_decorator(csrf_exempt, name='dispatch')
class CaseDataValidationView(viewsets.ViewSet):
    """
    Utility view for real-time data validation.
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def validate_case_id(self, request):
        """Validate case ID uniqueness and format"""
        case_id = request.data.get('case_id', '').strip().upper()
        current_case_id = request.data.get('current_case_id')  # For updates
        
        if not case_id:
            return Response({'valid': False, 'error': 'Case ID is required'})
        
        # Format validation
        if not re.match(r'^[A-Z0-9/\-\(\)\s\.]+$', case_id):
            return Response({
                'valid': False, 
                'error': 'Case ID can only contain uppercase letters, numbers, hyphens, forward slashes, parentheses, spaces, and dots'
            })
        
        # Uniqueness check
        queryset = Case.objects.filter(case_id__iexact=case_id)
        if current_case_id:
            queryset = queryset.exclude(case_id=current_case_id)
        
        if queryset.exists():
            return Response({'valid': False, 'error': 'A case with this Case ID already exists'})
        
        return Response({'valid': True})
    
    @action(detail=False, methods=['post'])
    def validate_mobile(self, request):
        """Validate 10-digit mobile number"""
        mobile = request.data.get('mobile', '').strip()
        
        if not mobile:
            return Response({'valid': True})  # Optional field
        
        # Remove spaces and hyphens
        mobile_clean = mobile.replace('-', '').replace(' ', '')
        
        # Validate Indian mobile number format
        if not re.match(r'^[6-9]\d{9}$', mobile_clean):
            return Response({
                'valid': False,
                'error': 'Mobile number must be exactly 10 digits starting with 6, 7, 8, or 9'
            })
        
        return Response({'valid': True, 'formatted': mobile_clean})
    
    @action(detail=False, methods=['post'])
    def format_currency(self, request):
        """Format currency in Indian format"""
        amount = request.data.get('amount', '')
        
        if not amount:
            return Response({'formatted': ''})
        
        try:
            # Parse amount
            if isinstance(amount, str):
                # Remove existing formatting
                clean_amount = re.sub(r'[^\d.]', '', amount)
                amount_decimal = Decimal(clean_amount)
            else:
                amount_decimal = Decimal(str(amount))
            
            # Format in Indian style
            formatted = self._format_indian_currency(amount_decimal)
            
            return Response({'formatted': formatted, 'valid': True})
            
        except (ValueError, TypeError):
            return Response({'formatted': '', 'valid': False, 'error': 'Invalid amount format'})
    
    def _format_indian_currency(self, amount):
        """Helper method for Indian currency formatting"""
        if not amount:
            return ''
        
        try:
            amount_str = f"{amount:,.2f}"
            parts = amount_str.split('.')
            integer_part = parts[0].replace(',', '')
            decimal_part = parts[1] if len(parts) > 1 else '00'
            
            if len(integer_part) > 3:
                result = integer_part[-3:]
                integer_part = integer_part[:-3]
                
                while len(integer_part) > 2:
                    result = integer_part[-2:] + ',' + result
                    integer_part = integer_part[:-2]
                
                if integer_part:
                    result = integer_part + ',' + result
            else:
                result = integer_part
            
            return f"Rs. {result}.{decimal_part}"
            
        except (ValueError, TypeError):
            return str(amount) if amount else ''
    @action(detail=False, methods=['post'])
    def validate_case_type(self, request):
        """Validate case type - allow any non-empty string"""
        case_type = request.data.get('case_type', '').strip()
    
        if not case_type:
            return Response({'valid': False, 'error': 'Case type is required'})
    
    # Basic validation - just ensure it's reasonable
        if len(case_type) > 50:
            return Response({'valid': False, 'error': 'Case type must be 50 characters or less'})
    
        return Response({'valid': True, 'normalized': case_type.upper()})

    @action(detail=False, methods=['post'])
    def validate_department(self, request):
        """Validate department - allow any non-empty string"""
        department = request.data.get('department', '').strip()
    
        if not department:
            return Response({'valid': False, 'error': 'Department is required'})
    
    # Basic validation
        if len(department) > 100:
            return Response({'valid': False, 'error': 'Department name must be 100 characters or less'})
    
        return Response({'valid': True, 'normalized': department})
